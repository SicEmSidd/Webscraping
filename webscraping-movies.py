
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font


#SOME USEFUL FUNCTIONS IN BEAUTIFULSOUP
#-----------------------------------------------#
# find(tag, attributes, recursive, text, keywords)
# findAll(tag, attributes, recursive, text, limit, keywords)

#Tags: find("h1","h2","h3", etc.)
#Attributes: find("span", {"class":{"green","red"}})
#Text: nameList = Objfind(text="the prince")
#Limit = find with limit of 1
#keyword: allText = Obj.find(id="title",class="text")
# wb = xl.Workbook()

# ws = wb.active

# ws.title = 'First Sheet'

# wb.create_sheet(index=1,title='Second Sheet')



# ws['A1'] = "Invoice"
# ws['A2'] = "Tired"
# ws['A3'] = "Brakes"
# ws['A4'] = "Alignment"
# header_font = Font(name = 'Times New Roman', size = 24, bold = True)
# ws['A1'].font = header_font

# ws.merge_cells('A1:B1')

#ws.unmerge_cells('A1:B1')

# ws['B2'] = 450
# ws['B3'] = 225
# ws['B4'] = 150

# ws['A8'] = 'Total'
# ws['A8'].font = Font(size = 16,bold = True)

# ws['B8'] = '=SUM(B2:B7)'

# ws.column_dimensions['A'].width = 25

# wb.save('PythonToExcel.xlsx')

#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2022/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

movie_rows = soup.findAll('tr')

for row in movie_rows[1:6]:
    td = row.findAll('td')
    print(td[0].text)
    print(td[1].text)
    print(td[5].text)
    print(td[6].text)

wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'Rank'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Gross'
ws['D1'] = 'Theaters'
ws['E1'] = 'Avg. Gross / Theater'

for x in range(1,6):
    td = movie_rows[x].findAll('td')
    rank = td[0].text
    title = td[1].text
    gross = int(td[5].text.replace('$','').replace(',',''))
    theaters = int(td[6].text.replace(',',''))

    avg = gross/theaters

    ws['A' + str(x+1)] = rank
    ws['B' + str(x+1)] = title
    ws['C' + str(x+1)] = gross
    ws['D' + str(x+1)] = theaters
    ws['E' + str(x+1)] = avg

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 25

header_font = Font(size = 16, bold = True)

for cell in ws[1:1]:
    cell.font = header_font
    

for cell in ws["D:D"]:
    cell.number_format = '#,##0'

for cell in ws["C:C"]:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws["E:E"]:
    cell.number_format = u'"$ "#,##0.00'

wb.save('BoxOfficeReport.xlsx')
##
##
##
##

